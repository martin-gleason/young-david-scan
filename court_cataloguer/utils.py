# File operations (USB detection, PDF import) and Excel export.
# No UI code here — pure logic only.

import ctypes
import os
import shutil
import string
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import database as db
from .config import ARCHIVE_DIR, EXPORTS_DIR, MASTER_XLSX
from .logging_setup import get_logger

log = get_logger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# USB DETECTION  (Windows only — uses ctypes, no extra packages)
# ══════════════════════════════════════════════════════════════════════════════

DRIVE_REMOVABLE = 2  # GetDriveTypeW return value for USB / removable media


def get_removable_drives() -> list[str]:
    """
    Return a list of removable drive root paths (e.g. ['E:\\', 'F:\\']).
    Uses the Windows API via ctypes — no external packages needed.
    Returns empty list on non-Windows or if no removable drives found.
    """
    if sys.platform != "win32":
        return []

    drives: list[str] = []
    try:
        bitmask = ctypes.windll.kernel32.GetLogicalDrives()
        for letter in string.ascii_uppercase:
            if bitmask & 1:
                drive = f"{letter}:\\"
                drive_type = ctypes.windll.kernel32.GetDriveTypeW(drive)
                if drive_type == DRIVE_REMOVABLE:
                    drives.append(drive)
            bitmask >>= 1
    except OSError:
        log.exception("Failed to enumerate logical drives")
    return drives


def find_pdfs_on_path(root: str) -> list[Path]:
    """
    Recursively find all PDF files under a root path.
    Skips folders it cannot read (PermissionError).
    """
    pdfs: list[Path] = []
    try:
        for dirpath, _, filenames in os.walk(root):
            for name in filenames:
                if name.lower().endswith(".pdf"):
                    pdfs.append(Path(dirpath) / name)
    except (PermissionError, OSError):
        log.exception("Failed walking %s", root)
    return pdfs


# ══════════════════════════════════════════════════════════════════════════════
# PDF IMPORT  (copy from USB → local archive, register in DB)
# ══════════════════════════════════════════════════════════════════════════════


def import_pdfs(pdf_paths: list[Path]) -> list[int]:
    """
    Copy each PDF into the local archive folder (organised by import date),
    register each one in the database as 'pending', and return a list of
    the new document IDs.

    Safe to call multiple times — duplicate filenames are auto-renamed.
    """
    today_str = datetime.now().strftime("%Y-%m-%d")
    dest_dir = ARCHIVE_DIR / today_str
    dest_dir.mkdir(parents=True, exist_ok=True)

    doc_ids: list[int] = []
    for src in pdf_paths:
        dest = _safe_copy(src, dest_dir)
        try:
            doc_id = db.add_document(
                original_filename=src.name,
                stored_path=str(dest),
                status="pending",
            )
        except Exception:
            # DB insert failed after the file copy succeeded — remove the
            # orphaned copy so the archive doesn't accumulate files we
            # don't know about.
            log.exception("add_document failed; removing orphaned copy %s", dest)
            try:
                dest.unlink(missing_ok=True)
            except OSError:
                log.exception("Failed to unlink orphan %s", dest)
            raise
        doc_ids.append(doc_id)
    return doc_ids


def _safe_copy(src: Path, dest_dir: Path) -> Path:
    """
    Copy src into dest_dir.  If a file with the same name already exists,
    append _v2, _v3, … until the name is unique.
    Returns the final destination path.
    """
    dest = dest_dir / src.name
    counter = 2
    while dest.exists():
        dest = dest_dir / f"{src.stem}_v{counter}{src.suffix}"
        counter += 1
    shutil.copy2(str(src), str(dest))
    return dest


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

# ── Style helpers ─────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
_HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill(start_color="EAF0F8", end_color="EAF0F8", fill_type="solid")
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_header(ws, headers: list[str]) -> None:
    ws.row_dimensions[1].height = 22
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=text)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _CENTER
        c.border = _BORDER


def _style_data_row(ws, row: int, n_cols: int) -> None:
    fill = _ALT_FILL if row % 2 == 0 else None
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.alignment = _LEFT
        c.border = _BORDER
        if fill:
            c.fill = fill


def _set_col_widths(ws, widths: list[int]) -> None:
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# ── Main export function ──────────────────────────────────────────────────────


def export_master_xlsx() -> Path:
    """
    Regenerate master_catalogue.xlsx from the current database contents.
    Overwrites the file if it already exists.
    Returns the path to the saved file.

    Raises PermissionError if the file is currently open in Excel.
    """
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # If the file is open in Excel, openpyxl fails late with a confusing error.
    # Probe by opening it for write — Excel takes an exclusive write lock.
    if MASTER_XLSX.exists():
        try:
            with open(MASTER_XLSX, "r+b"):
                pass
        except PermissionError as exc:
            raise PermissionError(
                "master_catalogue.xlsx is currently open in Excel.\n"
                "Please close it and try again."
            ) from exc

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Cases ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Cases"

    case_headers = [
        "Last Name",
        "Courtroom",
        "Docket #",
        "Case Date",
        "# Documents",
        "Date Created",
        "Notes",
    ]
    _write_header(ws1, case_headers)

    for r_idx, case in enumerate(db.get_all_cases(), start=2):
        doc_count = len(db.get_documents_for_case(case["id"]))
        values = [
            case["last_name"],
            case["courtroom"],
            case["docket_number"],
            case["case_date"],
            doc_count,
            case["created_at"],
            case.get("notes", ""),
        ]
        for c_idx, val in enumerate(values, start=1):
            ws1.cell(row=r_idx, column=c_idx, value=val)
        _style_data_row(ws1, r_idx, len(case_headers))

    _set_col_widths(ws1, [18, 16, 20, 14, 12, 22, 35])
    ws1.freeze_panes = "A2"
    total_cases = ws1.max_row
    if total_cases > 1:
        ws1.auto_filter.ref = f"A1:G{total_cases}"

    # ── Sheet 2: All Documents ─────────────────────────────────────────────
    ws2 = wb.create_sheet("All Documents")

    doc_headers = [
        "Docket #",
        "Last Name",
        "Courtroom",
        "Case Date",
        "Petition Type",
        "Filename",
        "Import Date",
        "Status",
    ]
    _write_header(ws2, doc_headers)

    for r_idx, doc in enumerate(db.get_all_documents(), start=2):
        values = [
            doc.get("docket_number", ""),
            doc.get("last_name", ""),
            doc.get("courtroom", ""),
            doc.get("case_date", ""),
            doc.get("petition_type", ""),
            doc.get("original_filename", ""),
            doc.get("imported_at", ""),
            doc.get("status", ""),
        ]
        for c_idx, val in enumerate(values, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=val)
        _style_data_row(ws2, r_idx, len(doc_headers))

    _set_col_widths(ws2, [20, 18, 16, 14, 38, 30, 22, 12])
    ws2.freeze_panes = "A2"
    total_docs = ws2.max_row
    if total_docs > 1:
        ws2.auto_filter.ref = f"A1:H{total_docs}"

    wb.save(str(MASTER_XLSX))
    return MASTER_XLSX
